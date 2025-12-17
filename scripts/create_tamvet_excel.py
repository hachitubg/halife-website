#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script tạo file Excel template cho quản lý sản phẩm Tâm Vet
Chạy: python scripts/create_tamvet_excel.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Dữ liệu mẫu
SAMPLE_PRODUCTS = [
    {
        'id': 'TAMVET001',
        'name': 'Thuốc Kháng Sinh Amoxicillin 500mg',
        'category': 'Thuốc Thú Cưng',
        'price': 45000,
        'originalPrice': 50000,
        'description': 'Thuốc kháng sinh hiệu quả điều trị các bệnh nhiễm khuẩn',
        'fullDescription': 'Amoxicillin 500mg là thuốc kháng sinh beta-lactam có hiệu quả cao trong điều trị các bệnh nhiễm khuẩn ở động vật.',
        'image': '/images/products/amoxicillin.jpg',
        'inStock': 'TRUE',
        'featured': 'TRUE',
        'packageSize': '10 vỉ/hộp',
        'stockQuantity': 150,
        'targetAnimal': 'Chó, Mèo, Gia súc',
        'manufacturer': 'Tâm Vet',
        'originCountry': 'Việt Nam',
        'registrationNumber': 'VN123456',
        'activeIngredients': 'Amoxicillin trihydrate 500mg',
        'dosage': 'Chó: 20-40mg/kg, Mèo: 15-25mg/kg',
        'usageInstructions': 'Uống sau ăn 2-3 lần/ngày',
        'warnings': 'Không sử dụng với thú có dị ứng penicillin',
        'storageConditions': 'Bảo quản nơi khô ráo, nhiệt độ 15-25°C',
        'rating': 4.5,
        'reviewCount': 32,
        'tags': 'kháng sinh;nhiễm khuẩn;chó;mèo',
        'functions': 'Kháng khuẩn;Diệt vi khuẩn'
    },
    {
        'id': 'TAMVET002',
        'name': 'Vitamin B Complex Cho Gia Súc',
        'category': 'Vitamin & Thực Phẩm Chức Năng',
        'price': 65000,
        'originalPrice': 75000,
        'description': 'Bổ sung vitamin B giúp tăng sức đề kháng cho gia súc',
        'fullDescription': 'Vitamin B Complex là một sản phẩm bổ sung vitamin nhóm B toàn diện, giúp cải thiện sức khỏe tổng thể.',
        'image': '/images/products/vitamin-b.jpg',
        'inStock': 'TRUE',
        'featured': 'FALSE',
        'packageSize': '100ml/chai',
        'stockQuantity': 200,
        'targetAnimal': 'Gia súc',
        'manufacturer': 'Tâm Vet',
        'originCountry': 'Việt Nam',
        'registrationNumber': 'VN123457',
        'activeIngredients': 'Thiamine,Riboflavin,Niacin,Pyridoxine',
        'dosage': '5-10ml/ngày/đầu gia súc',
        'usageInstructions': 'Trộn vào thức ăn hoặc cho uống trực tiếp',
        'warnings': 'Bảo quản trong nơi mát mẻ, tránh ánh sáng',
        'storageConditions': 'Bảo quản ở 2-8°C',
        'rating': 4.2,
        'reviewCount': 18,
        'tags': 'vitamin;bổ sung;gia súc',
        'functions': 'Tăng sức đề kháng;Cải thiện sức khỏe'
    },
    {
        'id': 'TAMVET003',
        'name': 'Chế Phẩm Vi Sinh Tăng Tiêu Hóa',
        'category': 'Chế Phẩm Sinh Học',
        'price': 85000,
        'originalPrice': 95000,
        'description': 'Chế phẩm vi sinh giúp cân bằng hệ tiêu hóa',
        'fullDescription': 'Chế phẩm này chứa các chủng vi sinh lợi ích giúp cân bằng hệ vi sinh đường ruột.',
        'image': '/images/products/probiotic.jpg',
        'inStock': 'TRUE',
        'featured': 'TRUE',
        'packageSize': '500g/kg',
        'stockQuantity': 120,
        'targetAnimal': 'Gia cầm, Gia súc, Cá',
        'manufacturer': 'Tâm Vet',
        'originCountry': 'Việt Nam',
        'registrationNumber': 'VN123458',
        'activeIngredients': 'Bacillus subtilis,Lactobacillus,Bifidobacterium',
        'dosage': '5-10g/tấn thức ăn',
        'usageInstructions': 'Trộn đều vào thức ăn tươi',
        'warnings': 'Bảo quản ở nhiệt độ thấp, tránh ẩm ướt',
        'storageConditions': 'Bảo quản ở 2-8°C hoặc -18°C',
        'rating': 4.8,
        'reviewCount': 45,
        'tags': 'vi sinh;tiêu hóa;chế phẩm',
        'functions': 'Tăng tiêu hóa;Cân bằng hệ vi sinh;Tăng trưởng'
    }
]

CATEGORIES = [
    {'id': 1, 'name': 'Thuốc Thú Cưng'},
    {'id': 2, 'name': 'Thuốc Gia Súc'},
    {'id': 3, 'name': 'Chế Phẩm Sinh Học'},
    {'id': 4, 'name': 'Vitamin & Thực Phẩm Chức Năng'},
    {'id': 5, 'name': 'Thiết Bị Y Tế'},
    {'id': 6, 'name': 'Gia Cầm'}
]

def create_template():
    """Tạo file Excel template"""
    
    # Tạo workbook
    wb = openpyxl.Workbook()
    
    # Xóa sheet mặc định
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ======== SHEET 1: Products ========
    ws_products = wb.create_sheet('Products', 0)
    
    # Header
    headers = list(SAMPLE_PRODUCTS[0].keys())
    for col_num, header in enumerate(headers, 1):
        cell = ws_products.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Dữ liệu
    for row_num, product in enumerate(SAMPLE_PRODUCTS, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws_products.cell(row=row_num, column=col_num)
            cell.value = product.get(header, '')
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Đặt độ rộng cột
            if col_num == 1:
                ws_products.column_dimensions[get_column_letter(col_num)].width = 12
            elif col_num in [2, 3, 8]:
                ws_products.column_dimensions[get_column_letter(col_num)].width = 20
            else:
                ws_products.column_dimensions[get_column_letter(col_num)].width = 18
    
    # Đặt độ cao hàng header
    ws_products.row_dimensions[1].height = 25
    
    # ======== SHEET 2: Categories ========
    ws_categories = wb.create_sheet('Categories', 1)
    
    # Header
    cat_headers = ['id', 'name']
    for col_num, header in enumerate(cat_headers, 1):
        cell = ws_categories.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dữ liệu
    for row_num, category in enumerate(CATEGORIES, 2):
        for col_num, header in enumerate(cat_headers, 1):
            cell = ws_categories.cell(row=row_num, column=col_num)
            cell.value = category.get(header, '')
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    ws_categories.column_dimensions['A'].width = 10
    ws_categories.column_dimensions['B'].width = 30
    ws_categories.row_dimensions[1].height = 25
    
    # ======== SHEET 3: Instructions ========
    ws_instructions = wb.create_sheet('Hướng Dẫn', 2)
    
    instructions = [
        ['HƯỚNG DẪN SỬ DỤNG FILE EXCEL'],
        [],
        ['SHEET "Products" (Sản phẩm):'],
        ['Các cột cần thiết (*)'],
        ['- id: Mã sản phẩm (bắt buộc)'],
        ['- name: Tên sản phẩm (bắt buộc)'],
        ['- category: Danh mục'],
        ['- price: Giá bán'],
        ['- description: Mô tả ngắn'],
        ['- inStock: TRUE/FALSE (Còn hàng)'],
        ['- featured: TRUE/FALSE (Sản phẩm nổi bật)'],
        [],
        ['SHEET "Categories" (Danh mục):'],
        ['- id: ID danh mục'],
        ['- name: Tên danh mục'],
        [],
        ['GHI CHÚ:'],
        ['1. Cột "id" phải có giá trị duy nhất'],
        ['2. Sử dụng TRUE hoặc FALSE (in hoa) cho giá trị boolean'],
        ['3. Tên sheet phải chính xác: "Products", "Categories"'],
        ['4. File phải lưu ở: public/data/tamvet_product.xlsx'],
        ['5. Định dạng: .xlsx (Excel 2007+)']
    ]
    
    for row_num, instruction in enumerate(instructions, 1):
        for col_num, value in enumerate(instruction, 1):
            cell = ws_instructions.cell(row=row_num, column=col_num)
            cell.value = value
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.font = Font(bold=True, size=14, color="FFFFFF")
            elif value.startswith('-') or value.startswith('1.') or value.startswith('2.') or value.startswith('3.') or value.startswith('4.') or value.startswith('5.'):
                cell.font = Font(bold=False)
            
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    ws_instructions.column_dimensions['A'].width = 50
    
    # Lưu file
    output_path = 'public/data/tamvet_product.xlsx'
    os.makedirs('public/data', exist_ok=True)
    wb.save(output_path)
    
    print(f'✅ Tạo file thành công: {output_path}')
    print(f'📊 File chứa:')
    print(f'   - Sheet "Products": {len(SAMPLE_PRODUCTS)} sản phẩm mẫu')
    print(f'   - Sheet "Categories": {len(CATEGORIES)} danh mục')
    print(f'   - Sheet "Hướng dẫn": Hướng dẫn sử dụng')

if __name__ == '__main__':
    try:
        create_template()
        print('\n💡 Bây giờ vào http://localhost:5173/admin/tamvet-products để kiểm tra!')
    except Exception as e:
        print(f'❌ Lỗi: {e}')
